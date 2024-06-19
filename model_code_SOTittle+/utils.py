import logging

import numpy as np
import pandas as pd
import os
import time
import torch
from io import open

from openpyxl import load_workbook, Workbook
from rouge import Rouge
from transformers import BertForSequenceClassification, BertTokenizer
from torch.utils.data import DataLoader, TensorDataset, RandomSampler, SequentialSampler
from openprompt import PromptDataLoader
from tqdm import tqdm
from openprompt.data_utils import InputExample
import nltk
import nltk.translate.gleu_score as gleu
from moverscore import get_idf_dict, word_mover_score
from bleurt.score import BleurtScorer
from bert_score import score as bert_score
from sentence_transformers import SentenceTransformer, util as sbert_util
import tensorflow_hub as hub
import tensorflow as tf

rouge = Rouge()



def get_elapse_time(t0):
    elapse_time = time.time() - t0
    if elapse_time > 3600:
        hour = int(elapse_time // 3600)
        minute = int((elapse_time % 3600) // 60)
        return "{}h{}m".format(hour, minute)
    else:
        minute = int((elapse_time % 3600) // 60)
        return "{}m".format(minute)


def convert_examples_to_features(examples, tokenizer, args, stage=None):
    # collect texts
    codes = []
    target_nl = []
    for example_id, example in enumerate(examples):
        codes.append(example.source)

        if stage == "test":
            target_nl.append("None")
        else:
            target_nl.append(example.target)

    # begin tokenizing
    encoded_codes = tokenizer(
        codes, padding=True, verbose=False, add_special_tokens=True,
        truncation=True, max_length=args.max_source_length, return_tensors='pt')

    encoded_targets = tokenizer(
        target_nl, padding=True, verbose=False, add_special_tokens=True,
        truncation=True, max_length=args.max_source_length, return_tensors='pt')

    return {'source_ids': encoded_codes['input_ids'], 'target_ids': encoded_targets['input_ids'],
            'source_mask': encoded_codes['attention_mask'], 'target_mask': encoded_targets['attention_mask']}

# Sentence-BERT
def get_sbert_score(references, hypotheses):
    model = SentenceTransformer('paraphrase-MiniLM-L6-v2')
    reference_embeddings = model.encode(references, convert_to_tensor=True)
    hypothesis_embeddings = model.encode(hypotheses, convert_to_tensor=True)
    cosine_scores = sbert_util.pytorch_cos_sim(reference_embeddings, hypothesis_embeddings)
    return cosine_scores.diagonal().mean().item()

# ROUGE
def get_rouge_score(references, hypotheses):
    rouge1_array = []
    rouge2_array = []
    rouge3_array = []
    for ref, hyp in zip(references, hypotheses):
        if isinstance(hyp, str) and isinstance(ref, str):
            if hyp.strip() == '' or ref.strip() == '':
                rouge1_array.append(0)
                rouge2_array.append(0)
                rouge3_array.append(0)
            else:
                rouge_scores = rouge.get_scores(hyp.strip(), ref.strip())
                rouge1_array.append(rouge_scores[0]["rouge-1"]['r'])
                rouge2_array.append(rouge_scores[0]["rouge-2"]['r'])
                rouge3_array.append(rouge_scores[0]["rouge-l"]['r'])
        else:
            rouge1_array.append(0)
            rouge2_array.append(0)
            rouge3_array.append(0)

    rouge_1 = np.mean(np.array(rouge1_array))
    rouge_2 = np.mean(np.array(rouge2_array))
    rouge_l = np.mean(np.array(rouge3_array))
    return rouge_1, rouge_2, rouge_l

def getAllRouge(preds, golden):
    with open(golden, 'r', encoding='utf8') as t, open(preds, 'r', encoding='utf8') as p:
        tline = t.readlines()
        pline = p.readlines()

    assert len(tline) == len(pline)

    rouge_1, rouge_2, rouge_l = get_rouge_score(tline, pline)
    sbert_score = get_sbert_score(tline, pline)

    ret_scores = {
        "ROUGE_1": rouge_1,
        "ROUGE_2": rouge_2,
        "ROUGE_L": rouge_l,
        "SBERT": sbert_score,
    }
    return ret_scores


def calculate_rouge(file_name, config, tokenizer, device, model, promptTemplate, WrapperClass,
                    output_file_name=None,
                    is_test=False, dev_dataloader=None,
                    best_rouge=None, lan=None):
    logging.basicConfig(format='%(asctime)s - %(levelname)s - %(name)s -   %(message)s',
                        datefmt='%m/%d/%Y %H:%M:%S',
                        level=logging.INFO)
    logger = logging.getLogger(__name__)
    logger.info("ROUGE file: {}".format(file_name))

    # whether append postfix to result file
    if output_file_name is not None:
        output_file_name = "_" + output_file_name
    else:
        output_file_name = ""

    if is_test:
        file_prefix = lan
    else:
        file_prefix = "dev"

    # if dev dataset has been saved
    if (not is_test) and (dev_dataloader is not None):
        eval_dataloader = dev_dataloader
    else:
        # read texts
        eval_examples = read_prompt_examples(file_name)

        # only use a part for dev
        # if not is_test:
        #     eval_examples = random.sample(eval_examples, min(1000, len(eval_examples)))

        eval_dataloader = PromptDataLoader(
            dataset=eval_examples,
            tokenizer=tokenizer,
            template=promptTemplate,
            tokenizer_wrapper_class=WrapperClass,
            max_seq_length=config.max_source_length,
            decoder_max_length=config.max_target_length,
            shuffle=False,
            teacher_forcing=False,
            predict_eos_token=True,
            batch_size=config.eval_batch_size,
        )

    model.eval()

    # generate texts by source
    generated_texts = []
    groundtruth_sentence = []
    guids = []
    for batch in tqdm(eval_dataloader, total=len(eval_dataloader)):
        batch = batch.to(device)
        with torch.no_grad():
            _, output_sentence = model.generate(batch, num_beams=10)
            generated_texts.extend(output_sentence)
            groundtruth_sentence.extend(batch['tgt_text'])
            guids.extend(batch['guid'])

    # compute rouge
    rouge_1, rouge_2, rouge_l = get_rouge_score(groundtruth_sentence,generated_texts)
    this_rouge = rouge_l

    if is_test:
        logger.info("  %s = %s " % ("ROUGE_L", str(this_rouge)))
    else:
        logger.info("  %s = %s \t Previous best ROUGE_L %s" % ("ROUGE_L", str(this_rouge), str(best_rouge)))

    logger.info("  " + "*" * 20)

    return this_rouge, eval_dataloader


def calculate_test(file_name, config, tokenizer, device, model, promptTemplate, WrapperClass,
                   output_file_name=None, lan=None):
    logging.basicConfig(format='%(asctime)s - %(levelname)s - %(name)s -   %(message)s',
                        datefmt='%m/%d/%Y %H:%M:%S',
                        level=logging.INFO)
    logger = logging.getLogger(__name__)
    logger.info("ROUGE file: {}".format(file_name))

    # whether append postfix to result file
    if output_file_name is not None:
        output_file_name = "_" + output_file_name
    else:
        output_file_name = ""

    file_prefix = lan

    # if dev dataset has been saved
    if True:
        # read texts
        eval_examples = read_prompt_examples(file_name)

        # only use a part for dev
        # if not is_test:
        #     eval_examples = random.sample(eval_examples, min(1000, len(eval_examples)))

        eval_dataloader = PromptDataLoader(
            dataset=eval_examples,
            tokenizer=tokenizer,
            template=promptTemplate,
            tokenizer_wrapper_class=WrapperClass,
            max_seq_length=config.max_source_length,
            decoder_max_length=config.max_target_length,
            shuffle=False,
            teacher_forcing=False,
            predict_eos_token=True,
            batch_size=config.eval_batch_size,
        )

        model.eval()

        # generate texts by source
        generated_texts = []
        groundtruth_sentence = []
        guids = []
        for batch in tqdm(eval_dataloader, total=len(eval_dataloader)):
            batch = batch.to(device)
            with torch.no_grad():
                _, output_sentence = model.generate(batch, num_beams=10)
                generated_texts.extend(output_sentence)
                groundtruth_sentence.extend(batch['tgt_text'])
                guids.extend(batch['guid'])

        # write to file
        with open(os.path.join( config.result_dir, file_prefix + ".pred.csv"), 'w',
                  encoding='utf-8') as f, \
            open(os.path.join(config.result_dir, file_prefix + ".gold.csv"), 'w',
            encoding='utf-8') as f1:

            for ref, gold, idx in zip(generated_texts, groundtruth_sentence, guids):
                f.write(ref + '\n')
                f1.write(gold + '\n')
        current_directory = r'{}'.format(os.path.dirname(os.path.abspath(__file__)))
    # compute rouge
    metrics_dict = getAllRouge(config.result_dir + r'{}.pred.csv'.format(
            file_prefix),
            config.result_dir + r'{}.gold.csv'.format(
            file_prefix))

    if not os.path.exists(config.result_dir + "evaluation.xlsx"):
        book = Workbook()
        sheet = book.active
        sheet.title = "MyApproach EVALUATION"
        fieldnames = ["Language", "Rouge-1", "Rouge-2", "Rouge-l", "SBERT", "USE"]
        sheet.append(fieldnames)
        data = [lan,
                f"{metrics_dict['ROUGE_1']}",
                f"{metrics_dict['ROUGE_2']}",
                f"{metrics_dict['ROUGE_L']}",
                f"{metrics_dict['SBERT']}"]

        sheet.append(data)
        book.save(config.result_dir + 'evaluation.xlsx')
    else:
        book = load_workbook(config.result_dir + 'evaluation.xlsx')
        sheet = book.active
        data = [lan,
                f"{metrics_dict['ROUGE_1']}",
                f"{metrics_dict['ROUGE_2']}",
                f"{metrics_dict['ROUGE_L']}",
                f"{metrics_dict['SBERT']}"]
        sheet.append(data)
        book.save(config.result_dir + 'evaluation.xlsx')

def read_prompt_examples(filename):
    """Read examples from filename."""
    examples = []
    data = pd.read_json(filename).astype(str)
    desc = data['desc'].tolist()  # 3 or desc
    code = data['code'].tolist()  # 2 or code
    title = data['title'].tolist()  # 0 or title
    tag = data['tags'].tolist()  # 1 or tags
    for idx in range(len(data)):
        examples.append(
            InputExample(
                guid=idx,
                text_a=desc[idx].lower(),
                text_b=tag[idx].lower() + ' : ' + code[idx].lower(),
                tgt_text=title[idx].lower(),
            )
        )

    return examples